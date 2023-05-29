import time
from openpyxl import load_workbook
import pandas as pd
from selenium.webdriver.common.by import By
from selenium import webdriver
import os.path
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from selenium_stealth import stealth

start_time = time.time()

path = r"E:\result25.xlsx" # The path to the file where the parsing results will be stored

def count_calls(func):
    """This decorator to track the number of times a function is called."""
    def wrapper(*args, **kwargs):
        wrapper.num_calls += 1
        return func(*args, **kwargs)
    wrapper.num_calls = 0
    return wrapper


def create_file():
    """This function checks if a file already exist. If the file doesn't exist, it creates a new
       Excel file by initializing a workbook and saving it at the specified path."""
    if not os.path.isfile(path):
        workbook = Workbook()
        workbook.save(path)


def get_main_data(row):
    """This function retrieves the main data from a LinkedIn company page, including the company name, location, and website link
       and then writes it to an Excel file. If the data is empty,
       it marks the corresponding cells in the Excel file with a specific color."""
    data = driver.execute_script("""
         data = {};
         const nolink = 'nolink';
         const empty = 'nothing';
         const maintag =  document.querySelector('.artdeco-entity-lockup--size-8'); 
            if(maintag){
               const companyName = maintag.querySelector('div.artdeco-entity-lockup__title').innerHTML;
               const location = maintag.querySelector('div.artdeco-entity-lockup__metadata div.t-black--light').innerHTML;
               data.company=companyName.trim();
               data.location = location.trim();
               const link = document.querySelector('a.view-website-link'); 
               if(link){
                 data.link = link.href;
                 console.log(data);
                 return data;
                 }else{
                    data.link = nolink;
                    console.log(data);
                    return data;
                 }     
              }else{
                return empty;
              }
              
           """)
    print('data:',data)
    print("writing the data to an Excel file...")
    my_file = load_workbook(filename=path)
    sheet = my_file.active
    row = row+1
    if data == 'nothing':
        sheet.cell(row=row, column=3).fill = PatternFill("solid", start_color="FAF205")
        sheet.cell(row=row, column=1).fill = PatternFill("solid", start_color="FAF205")
        sheet.cell(row=row, column=2).fill = PatternFill("solid", start_color="FAF205")
        my_file.save(path)
    else:
        sheet.cell(row=row, column=1, value=data['company'])
        sheet.cell(row=row, column=2, value=data['location'])
        if data['link'] == 'nolink':
            sheet.cell(row=row, column=1).fill = PatternFill("solid", start_color="FA8072")
            sheet.cell(row=row, column=2).fill = PatternFill("solid", start_color="FA8072")
            sheet.cell(row=row, column=3).fill = PatternFill("solid", start_color="FA8072")
        else:
            sheet.cell(row=row, column=3, value=data['link'])
        my_file.save(path)

    driver.close()
    time.sleep(2)
    driver.switch_to.window(driver.window_handles[0])



def go_to_link(links):
        """This function iterates through a list of links, opens each link in a new tab, and performs several actions,
           clicks the "Next" button to navigate to the next page of results. If the maximum number of iterations has not been reached, the function calls the
           "scroll_to_end_of_page" function again; otherwise, it quits the driver and exits the program."""
        for i, link in enumerate(links, start=1):
            print('*', i, link)
            time.sleep(5)
            excel = pd.read_excel(path, index_col=1)
            row = len(excel.index)+1
            driver.execute_script("""
                window.open(arguments[0], '_blank');
            """,link)
            time.sleep(5)
            driver.switch_to.window(driver.window_handles[-1])
            get_main_data(row)
            time.sleep(1)

        button = driver.execute_script("""
                   const button = document.querySelector('button.artdeco-pagination__button--next');
                   return button;
               """)

        driver.execute_script("arguments[0].click();", button)
        time.sleep(2)
        if not scroll_to_end_of_page.num_calls >= 2:
          scroll_to_end_of_page(driver)
        else:
            driver.quit()
            exit()

@count_calls
def scroll_to_end_of_page(driver):
    """This function scrolls to the end of a page in a LinkedIn"""
    print('--------------------')
    print('Page:', scroll_to_end_of_page.num_calls)
    print('--------------------')

    scroll_bar = driver.find_element_by_class_name("_vertical-scroll-results_1igybl")
    scroll_height = driver.execute_script("return arguments[0].scrollHeight", scroll_bar)
    scroll_top = 0
    scroll_step = 10
    global companies
    while scroll_top < scroll_height:
        driver.execute_script("arguments[0].scrollTop += arguments[1];", scroll_bar, scroll_step)
        companies = driver.execute_script("return document.querySelectorAll('.artdeco-list__item');" )
        time.sleep(0.1)
        scroll_top += scroll_step
    companies = list(set(companies))
    links = driver.execute_script("""
            const array = arguments[0];
            console.log('array: ',array);
            const links = [];
            array.forEach(div => {
              const aTags = div.querySelectorAll('.artdeco-entity-lockup__title a');
              aTags.forEach(link => links.push(link.href));
            });

            console.log(links);  
            return links

            """, companies)
    go_to_link(links)



def enter_to_linkedin():
    """This function performs the login to LinkedIn"""
    global driver
    opt = webdriver.ChromeOptions()
    opt.add_argument("--remote-debugging-port=9222")
    opt.add_argument("--no-sandbox")
    opt.add_argument("--disable-setuid-sandbox")
    opt.add_argument("disable-infobars")
    opt.add_experimental_option("excludeSwitches", ["enable-automation"])
    opt.add_experimental_option('useAutomationExtension', False)
    opt.add_argument('--disable-blink-features=AutomationControlled')
    driver = webdriver.Chrome(
        executable_path="C:\\Users\\josephbiden\\Desktop\\parser\\chromedriver\\chromedriver.exe",
        chrome_options=opt)
    stealth(
        driver,
        languages=["en-US", "en"],
        vendor="Google Inc.",
        platform="Win32",
        webgl_vendor="Intel Inc.",
        renderer="Intel Iris OpenGL Engine",
        fix_hairline=True,)
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": """
                    const newProto = navigator.__proto__
                    delete newProto.webdriver
                    navigator.__proto__ = newProto
                    """
    })
    driver.maximize_window()
    driver.get("https://www.linkedin.com/uas/login?session_redirect=/sales&fromSignIn=true&trk=navigator")
    time.sleep(10)
    driver.find_element(By.XPATH, '//*[@id="username"]').send_keys("login")
    time.sleep(15)
    driver.find_element(By.XPATH, '//*[@id="password"]').send_keys("password")
    time.sleep(10)
    button = driver.execute_script("""
                  let button = document.querySelector('.btn__primary--large');
                  console.log('button: ',button);
                  return button;
      """)
    button.click()
    time.sleep(30)
    driver.get("https://www.linkedin.com/sales/search/company?query=(filters%3AList((type%3AREGION%2Cvalues%3AList((id%3A102869081%2Ctext%3ABosnia%2520and%2520Herzegovina%2CselectionType%3AINCLUDED)%2C(id%3A106315325%2Ctext%3AGeorgia%2CselectionType%3AINCLUDED)%2C(id%3A105238872%2Ctext%3AIceland%2CselectionType%3AINCLUDED)%2C(id%3A102264497%2Ctext%3AUkraine%2CselectionType%3AINCLUDED)%2C(id%3A104640522%2Ctext%3AKosovo%2CselectionType%3AINCLUDED)%2C(id%3A100878084%2Ctext%3ALiechtenstein%2CselectionType%3AINCLUDED)%2C(id%3A104042105%2Ctext%3ALuxembourg%2CselectionType%3AINCLUDED)%2C(id%3A100961908%2Ctext%3AMalta%2CselectionType%3AINCLUDED)%2C(id%3A106178099%2Ctext%3AMoldova%2CselectionType%3AINCLUDED)%2C(id%3A101352147%2Ctext%3AMonaco%2CselectionType%3AINCLUDED)%2C(id%3A100733275%2Ctext%3AMontenegro%2CselectionType%3AINCLUDED)%2C(id%3A105730022%2Ctext%3ASan%2520Marino%2CselectionType%3AINCLUDED)%2C(id%3A101855366%2Ctext%3ASerbia%2CselectionType%3AINCLUDED)))%2C(type%3ACOMPANY_HEADCOUNT%2Cvalues%3AList((id%3AE%2Ctext%3A201-500%2CselectionType%3AINCLUDED)))%2C(type%3AINDUSTRY%2Cvalues%3AList((id%3A3132%2Ctext%3AInternet%2520Publishing%2CselectionType%3AINCLUDED)%2C(id%3A1285%2Ctext%3AInternet%2520Marketplace%2520Platforms%2CselectionType%3AINCLUDED)%2C(id%3A3131%2Ctext%3AMobile%2520Gaming%2520Apps%2CselectionType%3AINCLUDED)%2C(id%3A3100%2Ctext%3AMobile%2520Computing%2520Software%2520Products%2CselectionType%3AINCLUDED)%2C(id%3A3106%2Ctext%3AIT%2520System%2520Data%2520Services%2CselectionType%3AINCLUDED)%2C(id%3A3130%2Ctext%3AData%2520Security%2520Software%2520Products%2CselectionType%3AINCLUDED)%2C(id%3A5%2Ctext%3AComputer%2520Networking%2520Products%2CselectionType%3AINCLUDED)%2C(id%3A8%2Ctext%3ATelecommunications%2CselectionType%3AINCLUDED)%2C(id%3A3134%2Ctext%3ABlockchain%2520Services%2CselectionType%3AINCLUDED)))))&sessionId=15hBU9fJQvScfdO59M7Heg%3D%3D")
    time.sleep(5)
    create_file()
    scroll_to_end_of_page(driver)


if __name__ == "__main__":
    start_time = time.time()
    enter_to_linkedin()
    end_time = time.time()
    execution_time = end_time - start_time
    minutes = int(execution_time // 60)
    seconds = int(execution_time % 60)
    hours = minutes // 60
    minutes = minutes % 60

    print(f"Execution time: {hours} hours, {minutes} minutes")
